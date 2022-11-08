from openpyxl import load_workbook
import os

colunaPN = 2
colunaDescricao = 3
colunaTipo = 4
colunaUnidade = 6
colunaQuant = 6

itensDelete = ["90-000", "90-110", "4-260", "9-200", "4-220", "9-121",
               "9-135", "9-255", "9-536", "9-997", "10-251", "90-260"]


def formatChecklist(caminhoOriginal):
    # caminhoOriginal = "C:/Users/Roberto/Desktop/CHEKCLIST.xlsx"
    planilha = load_workbook(caminhoOriginal)
    folha = planilha['Planilha1']

    # print(f"Valor: {folha['A1'].value}")
    if deletColuns(folha):
        folha = deletColuns(folha)
        print("Colunas deletadas com sucesso!")
        folha = SelectDeletRows(folha)

        planilha.save(caminhoOriginal)
        planilha.close()
    else:
        print("Erro ao deletar colunas")


def SelectDeletRows(pagina):
    count = 0
    deletarLinhas = list()

    for linhas in pagina:
        count += 1
        for itemDelete in itensDelete:
            if itemDelete in linhas[colunaPN].value:
                # if itemDelete in "90-000":
                #      print(f"Valor da linha {count}: {linhas[colunaPN].value} --> Conjunto Montado Final")
                # elif itemDelete in "90-110":
                #      print("Conjunto Soldado")
                deletarLinhas.append(count)
            # else:
            #     print(f"Valor da linha {count}: {linhas[colunaPN].value}")
    print(f"Lista de linhas para serem excluidas: {deletarLinhas}")
    paginaTratada = deletRow(pagina, deletarLinhas)
    paginaFinal = organizarProduzidos(paginaTratada)
    return paginaFinal


def deletColuns(pagina):
    try:
        for index in range(8, 16):
            pagina.delete_cols(8)
        return pagina
    except:
        return False


def deletRow(pagina, linha):
    # print(f'Linhas recebidas para exclusao: {linha}')
    contador = 0
    for index in linha:
        # print(f'{index}  {type(index)}')
        pagina.delete_rows(index - contador)
        contador += 1

    # for item in pagina:
    #     print(item[colunaPN].value)

    return pagina


def organizarProduzidos(pagina):
    count = 0
    deletar = list()
    for linhas in pagina:
        # print(linhas[colunaPN].value)
        count += 1
        if "90-200" in linhas[colunaPN].value or "90-220" in linhas[colunaPN].value:
            pagina.cell(row=count, column=colunaUnidade).value = pagina.cell(row=count + 1,
                                                                             column=colunaDescricao + 1).value
            deletar.append(count + 1)
    deleteMP(pagina, deletar)


def deleteMP(pagina, deletar):
    contador = 0
    for index in deletar:
        pagina.delete_rows(index - contador)
        contador += 1


if __name__ == "__main__":
    formatChecklist()

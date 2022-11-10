from openpyxl import load_workbook, Workbook

path = "C:/Users/Roberto/Desktop/exemplo.xlsx"


tipo = 4
item = 2
quantidade = 6
descricao = 3


def format(caminhoOriginal, filtro):
    planilha = load_workbook(caminhoOriginal)
    sheets = planilha.sheetnames[0]
    folha = planilha[sheets]

    lista = selectRows(folha, filtro)
    dados = selectData(folha, lista)
    status = novaPlanilha(dados, 'C:/Users/Roberto/Desktop/exemplo_nova.xlsx')
    planilha.close()


def selectRows(folha, filtro) -> list:
    contador = 0
    linhas = list()
    for rows in folha:
        contador += 1
        if filtro in rows[item].value and 'P' in rows[tipo].value:
            # print(f'{rows[item].value}  {rows[tipo].value}  {rows[descricao].value}   {rows[quantidade].value}')
            linhas.append(contador)
    return linhas


def selectData(folha, lista) -> list:
    _item = 'C'
    _descricao = 'D'
    _tipo = 'E'
    _quantidade = 'G'
    contador = 0
    Colunas = []
    Matriz = []
    for row in folha:
        contador += 1
        indexMatriz = 0
        for index in lista:
            if contador == index:
                # print(f'{folha.cell(row=index, column=item + 1).value}  {row[tipo].value}')

                Colunas.append(folha.cell(row=index, column=item + 1).value)
                Colunas.append(folha.cell(row=index, column=descricao + 1).value)
                Colunas.append(folha.cell(row=index + 1, column=descricao + 1).value)
                Colunas.append(folha.cell(row=index, column=quantidade + 1).value)
                Colunas.append(folha.cell(row=index, column=tipo + 1).value)
                Matriz.append(Colunas.copy())
                Colunas.clear()
    #
    # for i in Matriz:
    #     print(i)

    return Matriz


def novaPlanilha(dados, path):
    wb = Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('Planilha1', 0)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 19
    ws.column_dimensions['C'].width = 23
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    cabecalho = ['ITEM', 'DESCRICAO', 'MATERIA PRIMA', 'QUANTIDADE', 'TIPO', ]
    ws.append(cabecalho)
    for row in dados:
        ws.append(row)
    wb.save(path)


if __name__ == "__main__":
    format(path, "90-200")
